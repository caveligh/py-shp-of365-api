from pathlib import PurePath, PureWindowsPath
from openpyxl import Workbook

import sys
import csv
import logging

sys.path.append('../')
from office365_api import SharePoint

# 1 args = The name of the SharePoint List
SHAREPOINT_LIST_NAME = sys.argv[1]
# 2 args = Export Type.  it can be "Excel" or "CSV"
EXPORT_TYPE = sys.argv[2]
# 3 args = Local Directory Path to save list dataset
DIR_PATH = sys.argv[3]
# 4 args = The name of the file that will get saved on local directory
FILE_NAME = sys.argv[4]

# Configuración de logs
logging.basicConfig(filename='log.txt', format='%(asctime)s , %(name)s , %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

def windows_path_to_wsl(windows_path):
    # Replaces backslashes with slashes
    path = windows_path.replace('\\', '/')

    # Divides the route in parts
    parts = path.split(':')

    # Replaces the drive letter with its equivalent in WSL and returns the new path
    # For example, the path 'C:\Users\principal\Downloads\files', would be translated in WSL as 
    # /mnt/c/Users/principal/Downloads/files.
    return '/mnt/{}/{}'.format(parts[0].lower(), parts[1])

def is_windows_path(path):
    return ':' in path or '\\' in path

def set_file_ext(file_name, export_type):
    if export_type == 'Excel':
        file_name_with_ext = '.'.join([file_name, 'xlsx'])
    elif export_type == 'CSV':
        file_name_with_ext = '.'.join([file_name, 'csv'])
    else:
        file_name_with_ext = file_name
    return file_name_with_ext

def download_list(list_name, export_type, dir_path, file_name):
    sp_list = SharePoint().get_list(list_name)
    if export_type == 'Excel':
        save_to_excel(sp_list, dir_path, file_name)
    elif export_type == 'CSV':
        save_to_csv(sp_list, dir_path, file_name)
    else:
        # Generamos mensaje de log
        logger.info(f"Export type is not a value type (download_list).")

def save_to_csv(list_items, dir_path, file_name):
    if is_windows_path(dir_path):
        dir_path = windows_path_to_wsl(dir_path)
    dir_file_path = PurePath(dir_path, file_name)

    # Generamos mensaje de log
    logger.info(f"Grabando archivo CSV {file_name} a la carpeta {dir_file_path}.")
    
    with open(dir_file_path, 'w', newline='\n', encoding='utf-8') as f:
        header = list_items[0].properties.keys()
        w = csv.DictWriter(f, header)
        w.writeheader()
        for item in list_items:
            w.writerow(item.properties)
            
def save_to_excel(list_items, dir_path, file_name):
    if is_windows_path(dir_path):
        dir_path = windows_path_to_wsl(dir_path)
    dir_file_path = PurePath(dir_path, file_name)
    wb = Workbook()
    ws = wb.active
    # Generamos mensaje de log
    logger.info(f"Grabando archivo excel {file_name} a la carpeta {dir_file_path}.")
    # list of header name from SharePoint List
    header = list_items[0].properties.keys()
    # write headers on first row
    for idx, name in enumerate(header):
        ws.cell(row=1, column=idx + 1, value=name)
    # write line items starting on second row
    row = 2
    for dict_obj in list_items:
        for idx, item in enumerate(dict_obj.properties.items()):
            ws.cell(row=row, column=idx + 1, value=item[1])
        row += 1
    wb.save(dir_file_path)
    
if __name__ == '__main__':
    try:
        file_name = set_file_ext(FILE_NAME, EXPORT_TYPE)
        download_list(SHAREPOINT_LIST_NAME, EXPORT_TYPE, DIR_PATH, file_name)
    except Exception as e:
        # Generamos mensaje de log si ocurre un error durante la ejecución del script
        logger.exception("Ha ocurrido un error durante la ejecución del método download_list.")
    finally:
        # Cerramos el archivo de logs
        logging.shutdown()  
            